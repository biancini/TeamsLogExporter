from os import chdir
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import locale

locale.setlocale(locale.LC_TIME, "it_IT")

nome_excel = 'Attivazioni Office 365 classi 1'
excel_path = f'/Users/andrea/Fondazione Enaip Lombardia/Istruzione e Formazione Professionale - Anno Scolastico 2020 2021/Attivazioni Office 365/'
base = '/Users/andrea/Downloads/Nuove utenze.xlsx'
chdir(excel_path)

centri = {
    'Bergamo': b'Bergamo',
    'Busto Arsizio': b'Busto',
    'Cantù': b'Cant\xc3\xb9',
    'Como': b'Como',
    'Cremona': b'Cremona',
    'Dalmine': b'Dalmine',
    'Lecco': b'Lecco',
    'Magenta': b'Magenta',
    'Mantova': b'Mantova',
    'Melzo': b'Melzo',
    'Milano Giacinti': b'Milano',
    'Monticello': b'Monticello',
    'Morbegno': b'Morbegno',
    'Pavia': b'Pavia',
    'Romano': b'Romano',
    'Varese': b'Varese',
    'Vigevano': b'Vigevano',
    'Vimercate': b'Vimercate',
    'Voghera': b'Voghera'
}


wb = load_workbook(filename=base)
ws = wb['studenti']

for c, cc in centri.items():
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = 'studenti'

    count = 0
    for row in ws.iter_rows():
        sede = row[0]
        if sede.value is None:
            break
        sede = sede.value.encode("utf8")
        if sede == b'Sede ENAIP':
            new_ws.append([r.value for r in row])
            continue

        if sede != cc:
            continue

        new_ws.append([r.value for r in row])
        count += 1

    mediumStyle = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    if count == 0:
        new_ws.add_table(Table(ref=f'A1:L{count+2}', displayName='Studenti', tableStyleInfo=mediumStyle))
    else:
        new_ws.add_table(Table(ref=f'A1:L{count+1}', displayName='Studenti', tableStyleInfo=mediumStyle))

    newpath = f'{nome_excel} - {c}.xlsx'
    new_wb.save(filename=newpath)
    new_wb.close()

    print(f'Generato il file per {c} con {count} iscritti')

print("Script finito.")