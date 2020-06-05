from os import path, chdir, getcwd
from glob import glob
from datetime import datetime, timedelta
from openpyxl import load_workbook
from tqdm import tqdm
import csv
import warnings

warnings.simplefilter("ignore")

start_date = datetime(2020, 4, 1)
end_date = datetime.now()

stats = {}

if __name__ == '__main__':
    cwd = getcwd()
    base = 'D:\Fondazione Enaip Lombardia\Pianificazione Attività - Documenti'
    chdir(base)
    files = glob(f'./**/*.xlsx', recursive=True)

    for f in tqdm(files):
        if '00_Generale' in f: continue

        try:
            file_date = datetime.strptime(path.basename(f)[:10], '%Y-%m-%d')
            if file_date < start_date: continue 
            if file_date > end_date: continue
        except ValueError:
            continue

        file_date = file_date.strftime('%d/%m/%Y')

        wb = load_workbook(f, read_only=True)
        ws = wb.active

        if not file_date in stats.keys():
            stats[file_date] = {
                'students': [],
                'hours': 0,
                'lessons': 0
            }

        hour_count = 0
        for row in ws.iter_rows(min_row=2, max_col=4, max_row=1000, values_only=True):
            if row[0] is not None and row[0] != '':
                stats[file_date]['students'].append(row[0])

                if type(row[1]) is datetime and type(row[2]) is datetime:
                    delta = (row[2] - row[1]).seconds
                    delta /= 60*60
                    if delta > hour_count:
                        hour_count = delta

        stats[file_date]['hours'] += int(hour_count)
        stats[file_date]['lessons'] += 1

        wb._archive.close()

    chdir(cwd)
    with open('stats.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Giorno", "Numero studenti", "Numero ore", "Numero lezioni"])

        for key, value in stats.items():
            writer.writerow([key, len(set(value['students'])), value['hours'], value['lessons']])

    print(f'Script finito.')