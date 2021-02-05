from datetime import datetime, timedelta
from os import chdir
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import locale

locale.setlocale(locale.LC_TIME, "it_IT")

excel_file = '/Users/andrea/Downloads/Iscrizioni corsi IeFP - Istruzione e Formazione Professionale_Anno Formativo 2020_2021.xlsx'
#base = 'D:\Fondazione Enaip Lombardia\Pianificazione Attività - Documenti'
base = '/Users/andrea/Fondazione Enaip Lombardia/Istruzione e Formazione Professionale - Anno Scolastico 2020 2021/Iscrizioni'
chdir(base)

nome_excel = 'Iscritti 2-8nov 20.xlsx'
data_da = datetime(2020, 11, 2)
data_a = datetime(2020, 11, 8)

centri = {
    'Bergamo': b'Bergamo',
    'Busto Arsizio': b'Busto Arsizio',
    'Cantù': b'Cant\xc3\xb9',
    'Como': b'Como',
    'Cremona': b'Cremona',
    'Dalmine': b'Dalmine',
    'Lecco': b'Lecco',
    'Magenta': b'Magenta',
    'Mantova': b'Mantova',
    'Melzo': b'Melzo',
    'Milano Giacinti': b'Milano Giacinti',
    'Monticello': b'Monticello',
    'Morbegno': b'Morbegno',
    'Pavia': b'Pavia',
    'Romano': b'Romano',
    'Varese': b'Varese',
    'Vigevano': b'Vigevano',
    'Vimercate': b'Vimercate',
    'Voghera': b'Voghera'
}

wb = load_workbook(filename=excel_file)
ws = wb['Form1']

for c, cc in centri.items():
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = 'Form1'

    count = 0
    for row in ws.iter_rows():
        data_iscrizione = row[2].value
        sede = row[26]
        if sede.value is None:
            continue
        sede = row[26].value.encode("utf8")

        if data_iscrizione == 'Ora di completamento':
            new_ws.append([r.value for r in row])
            continue

        if sede != cc:
            continue

        if not isinstance(data_iscrizione, datetime):
            continue
        
        if data_iscrizione < data_da or data_iscrizione > data_a + timedelta(days=1):
            continue
        
        new_ws.append([r.value for r in row])
        count += 1

    
    mediumStyle = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    if count == 0:
        new_ws.add_table(Table(ref=f'A1:DA{count+2}', displayName='RegistroPresenze', tableStyleInfo=mediumStyle))
    else:
        new_ws.add_table(Table(ref=f'A1:DA{count+1}', displayName='RegistroPresenze', tableStyleInfo=mediumStyle))

    newpath = f'{base}/Schede Iscrizione/{c}/{nome_excel}'
    new_wb.save(filename=newpath)
    new_wb.close()

    print(f'Generato il file per {c} con {count} iscritti')

print("Script finito.")