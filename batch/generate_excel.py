import json
import glob
import os.path
import sys
import getopt
import configparser
import logging
import requests
from tqdm import tqdm
from dateutil import tz
from concurrent.futures import ProcessPoolExecutor
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from utils import get_access_token

usernames = {}
from_zone = tz.gettz('UTC')
to_zone = tz.gettz('Europe/Rome')

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger()


def get_usernamefromid(t, userid, displayName=False):
    if userid == 'Sconosciuto': return 'Sconosciuto'

    keyname = userid
    username = f'Sconosciuto ({userid})'
    if displayName:
        keyname = f'{keyname}_d'

    if keyname not in usernames:
        uri = f'https://graph.microsoft.com/beta/users/{userid}'
        head = { 'Authorization': f'Bearer {t}' }
        r = requests.get(uri, headers=head)
        user = r.json()

        if displayName:
            if 'displayName' in user:
                usernames[keyname] = user['displayName']
            else:
                usernames[keyname] = username
        else:   
            if 'surname' in user and 'givenName' in user:            
                usernames[keyname] = f'{user["surname"]} {user["givenName"]}'
            else:
                usernames[keyname] = username

    return usernames[keyname]


def sheet_registro(filename, participants, workbook):
    worksheet = workbook.active
    worksheet.title = 'Registro'
    worksheet.append(['Partecipante', 'Inizio presenza', 'Fine presenza', 'Tempo di partecipazione'])
    for cell in worksheet["1:1"]:
        cell.font = Font(bold=True)
    i = 1
    for pp in participants:
        i = i + 1
        min_start = datetime.strptime(pp['start'].split('.', 1)[0], '%Y-%m-%dT%H:%M:%S')
        max_end = datetime.strptime(pp['end'].split('.', 1)[0], '%Y-%m-%dT%H:%M:%S')
        duration = 0
        for [start, end] in merge_intervals(pp['periods']):
            delta = end - start
            duration += delta.seconds
        duration /= 60*60*24
        worksheet.append([
                pp['name'],
                min_start,
                max_end,
                duration
            ])
        cell = worksheet.cell(i, 2)
        cell.number_format = 'dd/mm/yyyy hh:mm'
        cell = worksheet.cell(i, 3)
        cell.number_format = 'dd/mm/yyyy hh:mm'
        cell = worksheet.cell(i, 4)
        cell.number_format = 'h "ore e" mm "minuti"'
    mediumStyle = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    worksheet.add_table(Table(ref=f'A1:D{i}', displayName='RegistroPresenze', tableStyleInfo=mediumStyle))
    worksheet.sheet_view.showGridLines = False
    report_id = os.path.basename(filename).replace('.json', '')
    worksheet.append([f''])
    worksheet.append([f'Report generato per il meeting con ID: {report_id}'])
    column_widths = [30, 20, 20, 20]
    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i+1)].width = column_width


def merge_intervals(temp_tuple):
    temp_tuple.sort(key=lambda interval: interval[0])
    merged = [temp_tuple[0]]
    for current in temp_tuple:
        previous = merged[-1]
        if current[0] <= previous[1]:
            previous[1] = max(previous[1], current[1])
        else:
            merged.append(current)
    return merged


def sheet_partecipazione(filename, start_time, participants, workbook):
    worksheet = workbook.create_sheet('Partecipazione')
    titles = ['Partecipante']
    max_interruptions = 0
    for pp in participants:
        if len(pp['periods']) > max_interruptions:
            max_interruptions = len(pp['periods'])

    for i in range(max_interruptions):
        titles.append(f'Inizio {i+1}')
        titles.append(f'Fine {i+1}') 

    worksheet.append(titles)
    for cell in worksheet["1:1"]:
        cell.font = Font(bold=True)
    
    i = 1
    for pp in participants:
        i = i + 1

        cur_start = start_time
        row = [pp['name']]
        for j in range(len(pp['periods'])):
            p = pp['periods'][j]
            row.append(p[0] - cur_start)
            row.append(p[1] - p[0])
            cur_start = p[1]

        worksheet.append(row)
        for ii in range(1, len(row)+1):
            cell = worksheet.cell(i, ii)
            cell.number_format = 'hh "ore e" mm "minuti"'

    column_widths = [30]
    for _i in range(max_interruptions*2):
        column_widths.append(20)
    for ii, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(ii+1)].width = column_width
        
    max_col = get_column_letter((max_interruptions*2)+1)
    mediumStyle = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    worksheet.add_table(Table(ref=f'A1:{max_col}{i}', displayName='RegistroPartecipazione', tableStyleInfo=mediumStyle))
    worksheet.sheet_view.showGridLines = False
    report_id = os.path.basename(filename).replace('.json', '')
    worksheet.append([f''])
    worksheet.append([f'Report generato per il meeting con ID: {report_id}'])

    data = Reference(worksheet, min_col=2, min_row=1, max_row=i, max_col=(max_interruptions*2)+1)
    cats = Reference(worksheet, min_col=1, min_row=2, max_row=i)
    chart = BarChart()
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.legend = None 
    chart.height = 20
    chart.width = 40
    chart.x_axis.scaling.orientation = "maxMin"
    for ii, s in enumerate(chart.series):
        s.graphicalProperties.line.nofill = True
        s.graphicalProperties.line.shadow = None

        if ii % 2 == 0:
            s.graphicalProperties.noFill = True
        else:
            s.graphicalProperties.solidFill = "008F00"

    worksheet.add_chart(chart, "A43")


def generate_one_excel(t, filename):
    with open(filename) as json_file:
        p = json.load(json_file)

        name = 'Sconosciuto'
        if 'organizer' in p and p['organizer'] is not None:
            if 'user' in p['organizer'] and p['organizer']['user'] is not None:
                if 'id' in p['organizer']['user']:
                    name = get_usernamefromid(t, p['organizer']['user']['id'], True)

        start_time = datetime.strptime(p['startDateTime'].split('.', 1)[0].split('Z', 1)[0], '%Y-%m-%dT%H:%M:%S')
        start_time = start_time.replace(tzinfo=from_zone).astimezone(to_zone)
        start_time_name = start_time.strftime('%Y-%m-%d %H-%M')
        dest_filename = f'excel/{start_time_name} - {name}.xlsx'

        i = 2
        while os.path.isfile(dest_filename):
            dest_filename = f'excel/{start_time_name} - {name} - {i}.xlsx'
            i = i + 1

        users = {}
        for c in p['sessions']:
            if 'caller' not in c or c['caller'] is None:
                continue
            if 'identity' not in c['caller'] or c['caller']['identity'] is None:
                continue
            if 'user' not in c['caller']['identity'] or c['caller']['identity']['user'] is None:
                continue
            if 'id' not in c['caller']['identity']['user'] or c['caller']['identity']['user']['id'] is None:
                continue

            if 'userAgent' in c['caller'] and 'headerValue' in c['caller']['userAgent'] and c['caller']['userAgent']['headerValue'] == 'SkypeBot Call Recorder Teams':
                continue

            curuid = c['caller']['identity']['user']['id']
            if curuid not in users:
                displayname = get_usernamefromid(t, curuid)
                users[curuid] = { 'name': displayname, 'min_start': None, 'max_end': None }

            if 'periods' not in users[curuid]: users[curuid]['periods'] = []
            
            start = datetime.strptime(c['startDateTime'].split('.', 1)[0].split('Z', 1)[0], '%Y-%m-%dT%H:%M:%S')
            start = start.replace(tzinfo=from_zone).astimezone(to_zone)

            if users[curuid]['min_start'] is None or start < users[curuid]['min_start']:
                users[curuid]['min_start'] = start

            end = datetime.strptime(c['endDateTime'].split('.', 1)[0].split('Z', 1)[0], '%Y-%m-%dT%H:%M:%S')
            end = end.replace(tzinfo=from_zone).astimezone(to_zone)

            if users[curuid]['max_end'] is None or end > users[curuid]['max_end']:
                users[curuid]['max_end'] = end
            
            users[curuid]['periods'].append([start, end])

        participants = []
        for _uid, data in users.items():
            participants.append({
                'uid': _uid,
                'name': data['name'],
                'start': data['min_start'].strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
                'end': data['max_end'].strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
                'periods': merge_intervals(data['periods'])})

        json_file.close()
        if len(participants) <= 1:
            return [{'file': dest_filename, 'generated': False, 'message': 'Only one participant'}]

    participants.sort(key=lambda stud: 'zzz_{}'.format(stud['name']) if 'Sconosciuto' in stud['name'] else stud['name'] )

    workbook = Workbook()
    sheet_registro(filename, participants, workbook)
    sheet_partecipazione(filename, start_time, participants, workbook)
    workbook.save(filename=dest_filename)

    return [{'file': dest_filename, 'generated': True}]


def generate_excel(configuration):
    ente = configuration['ente']
    filename = None

    t = get_access_token(ente)
    json_files = glob.glob("json/*.json")
    out = 0
    num_threads = 10

    with ProcessPoolExecutor(max_workers=num_threads) as pool:
        with tqdm(total=len(json_files)) as progress:
            futures = []
            for filename in json_files:
                future = pool.submit(generate_one_excel, t, filename)
                futures.append(future)

            for future in futures:
                progress.update()
                result = future.result()
                out += len([r for r in result if r['generated']]) if result is not None else 0
                for r in result:
                    if r['generated']:
                        logging.debug("(%d/%d) Generated excel file: %s", out, len(json_files), r['file'])
                    else:
                        logging.debug("(%d/%d) Not generated excel file: %s, with message %s", out, len(json_files), r['file'], r['message'])
        
    return out
    

if __name__ == '__main__':
    config = configparser.ConfigParser()
    config.read('configuration.ini')
    ente = 'ENAIP'

    try:
        opts, _ = getopt.getopt(sys.argv[1:], "he:", ["help", "ente="])
    except getopt.GetoptError:
        print('generate_excel.py [-e <ente>]')
        sys.exit(2)
    
    for o, a in opts:
        if o in ('-h', '--help'):
            print('generate_excel.py [-e <ente>]')
            sys.exit()
        elif o in ('-e', '--ente'):
            ente = a.upper()
        else:
            assert False

    configuration = config[ente]
    configuration['ente'] = ente

    print(f'Working for institution {ente}.')
    numexcels = generate_excel(configuration)
    print(f'Script finito, creati {numexcels} files.')
