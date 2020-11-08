import json
import math
import os
import urllib.parse
import traceback
from io import StringIO
from dateutil import tz
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, time
from exporter.auth_helper import scopes, get_sign_in_url, get_token_from_code, store_token, store_bearertoken, store_user, remove_user_and_token, get_user, get_token, get_bearertoken
from exporter.graph_helper import get_meuser, get_otheruser, get_all_users, get_all_groups, get_group_users, get_user_meetings, get_meeting_records


from_zone = tz.gettz('UTC')
to_zone = tz.gettz('Europe/Rome')

def initialize_context(request):
    context = {}

    # Check for any errors in the session
    error = request.session.pop('flash_error', None)

    if error is not None:
        context['errors'] = []
        context['errors'].append(error)

    # Check for user in the session
    context['user'] = request.session.get('user', {'is_authenticated': False})
    return context

def sign_in(request):
    id = request.GET.get('id', None)
    request.session['appid'] = os.environ['APPID_%s' % id]
    request.session['appsecret'] = os.environ['APPSECRET_%s' % id]
    request.session['ente'] = id

    print ("appid = %s" % request.session.get('appid', None))
    print ("appsecret = %s" % request.session.get('appsecret', None))

    # Get the sign-in URL
    sign_in_url, state = get_sign_in_url(request)
    # Save the expected state so we can validate in the callback
    request.session['auth_state'] = state
    # Redirect to the Azure sign-in page
    return HttpResponseRedirect(sign_in_url)

def callback(request):
    # Get the state saved in session
    expected_state = request.session.pop('auth_state', '')
    # Make the token request
    token = get_token_from_code(request, request.get_full_path(), expected_state)

    # Get the user's profile
    user = get_meuser(token)

    # Save token and user
    store_token(request, token)
    store_user(request, user)

    return HttpResponseRedirect(reverse('home'))

def sign_out(request):
    # Clear out the user and token
    remove_user_and_token(request)

    return HttpResponseRedirect(reverse('home'))

def home(request):
    context = initialize_context(request)
    context['appdata'] = {
        'appid': request.session.get('appid', None),
        'ente': request.session.get('ente', None),
        'scopes': scopes
    }

    return render(request, 'exporter/home.html', context)

def bearer(request):
    message = "Effettuata la chiamata in modo errato."

    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            if 'token' in data and data['token']:
                store_bearertoken(request, data['token'])

            token = get_token(request)
            groups = get_all_groups(token)

            data = [{ 'name': g['displayName'], 'id': g['id'] } for g in groups]
            data = sorted(data, key=lambda k: k['name'])

            return JsonResponse({
                "esito": True,
                "message": "Registrato bearer, ecco la lista dei gruppi.",
                "data": data
            })
    except Exception as e:
        message = "%s" % e
        print("Error: %s" % message)
        
    return JsonResponse({
        "esito": False,
        "message": message
    }, status=500)

def getusers_bygroup(request):
    message = "Effettuata la chiamata in modo errato."

    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            group_ids = data['groups']

            token = get_token(request)
            
            if len(group_ids) > 0:
                users = []
                for gid in group_ids:
                    newusers = get_group_users(token, gid)
                    users.extend(newusers)
            else:
                users = get_all_users(token)

            data = [{ 'name': u['displayName'], 'id': u['id'] } for u in users]
            data = sorted(data, key=lambda k: k['name'])

            return JsonResponse({
                "esito": True,
                "message": "Ecco la lista degli utenti.",
                "data": data
            })
    except Exception as e:
        message = "%s" % e
        print("Error: %s" % message)
        
    return JsonResponse({
        "esito": False,
        "message": message
    }, status=500)

def getuser_meetings(request):
    message = "Effettuata la chiamata in modo errato."

    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            user_ids = data['users']

            token = get_bearertoken(request)

            events = []
            if len(user_ids) > 0:
                for uid in user_ids:
                    newevents = get_user_meetings(token, uid)
                    events.extend(newevents)

            data = [{ 'start': e['startDateTime'], 'end': e['endDateTime'], 'partecipant': e['participantCount'], 'id': e['id'] } for e in events]
            data = sorted(data, key=lambda k: k['start'])

            return JsonResponse({
                "esito": True,
                "message": "Ecco la lista degli utenti.",
                "data": data
            })
    except Exception as e:
        message = "%s" % e
        print("Error: %s" % message)

    return JsonResponse({
        "esito": False,
        "message": message
    }, status=500)

def getmeeting_records(request):
    message = "Effettuata la chiamata in modo errato."

    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            event_ids = data['meetings']

            token = get_bearertoken(request)
            graph_token = get_token(request)
            uid = get_user(request)['id']
            
            meeting_records = []
            
            if len(event_ids) > 0:
                for eid in event_ids:
                    result = get_meeting_records(token, uid, eid)
                    participants = []

                    for p in result:
                        uid = p['userId']
                        user = get_otheruser(graph_token, uid)
                        name = user['displayName'] if 'displayName' in user else 'Sconosciuto'

                        min_start = None
                        max_end = None
                        duration = 0

                        for c in p['communicationFragments']:
                            start = datetime.strptime(e['startDateTime'].split('.', 1)[0].split('Z', 1)[0], '%Y-%m-%dT%H:%M:%S')
                            start = start.replace(tzinfo=from_zone).astimezone(to_zone)
                            start = start.strftime('%Y-%m-%d %H-%M')

                            if min_start is None or start < min_start:
                                min_start = start

                            end = datetime.strptime(e['endDateTime'].split('.', 1)[0].split('Z', 1)[0], '%Y-%m-%dT%H:%M:%S')
                            end = end.replace(tzinfo=from_zone).astimezone(to_zone)
                            end = end.strftime('%Y-%m-%d %H-%M')
                            if max_end is None or end > max_end:
                                max_end = end
                            
                        
                            if end is not None and start is not None:
                                delta = max_end - min_start
                                duration += delta.seconds
                            
                        duration /= 60
                        hours = math.floor(duration / 60)
                        minutes = math.floor(duration % 60)
                        duration = "{0} ore e {1} minuti".format(hours, minutes)

                        participants.append({ 'uid': uid, 'name': name, 'start': min_start.strftime('%Y-%m-%dT%H:%M:%S.%fZ'), 'end': max_end.strftime('%Y-%m-%dT%H:%M:%S.%fZ'), 'duration': duration})
                        participants = sorted(participants, key=lambda k: k['name'])

                    meeting_records.append({ 'id': eid, 'participants': participants })

            return JsonResponse({
                "esito": True,
                "message": "Ecco la lista degli utenti.",
                "data": meeting_records
            })

    except Exception as e:
        message = "%s" % e
        print("Error: %s" % message)
        traceback.print_exc()
        
    return JsonResponse({
        "esito": False,
        "message": message
    }, status=500)

def export_xls(request):
    message = "Effettuata la chiamata in modo errato."

    try:
        if request.method == 'POST':
            data = urllib.parse.unquote(request.POST.get("table", '\{\}'))
            table = json.loads(data)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Registro'

            worksheet.append(['Partecipante', 'Inizio presenza', 'Fine presenza', 'Tempo di partecipazione'])
            for cell in worksheet["1:1"]:
                cell.font = Font(bold=True)

            i = 1
            for t in table['participants']:
                i = i + 1

                duration = t['duration'].replace(" ore e ", ":").replace(" minuti", "").split(":")
                worksheet.append([
                    t['name'],
                    datetime.strptime(t['start'].split('.', 1)[0], '%Y-%m-%dT%H:%M:%S'),
                    datetime.strptime(t['end'].split('.', 1)[0], '%Y-%m-%dT%H:%M:%S'),
                    time(int(duration[0]), int(duration[1]), 0)
                ])

                cell = worksheet.cell(i, 2)
                cell.number_format = 'dd/mm/yyyy hh:mm'
                cell = worksheet.cell(i, 3)
                cell.number_format = 'dd/mm/yyyy hh:mm' 
                cell = worksheet.cell(i, 4)
                cell.number_format = 'h "ore e" mm "minuti"' 

            mediumStyle = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
            worksheet.add_table(Table(ref='A1:D%s'%i, displayName='RegistroPresenze', tableStyleInfo=mediumStyle))
            worksheet.sheet_view.showGridLines = False

            column_widths = [30, 20, 20, 20]
            for i, column_width in enumerate(column_widths):
                worksheet.column_dimensions[get_column_letter(i+1)].width = column_width

            response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="{0}.xlsx"'.format(table['descr'].replace(':', 'h'))
            return response

    except Exception as e:
        message = "%s" % e
        print("Error: %s" % message)
    
    return JsonResponse({
        "esito": False,
        "message": message
    }, status=500)