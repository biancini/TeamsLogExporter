import requests
import json
import math
import os.path
from urllib import parse
from dateutil import tz
from urllib import parse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from io import BytesIO
from tempfile import NamedTemporaryFile

'''
Scaricare i dati dal report creato appostivamente qui:
https://cqd.teams.microsoft.com/spd/#/Dashboard?language=it-IT
'''

from_zone = tz.gettz('UTC')
to_zone = tz.gettz('Europe/Rome')

def get_berarertoken(ente):
    tenant_id = os.environ[f'TENANTID_{ente}']
    client_id = os.environ[f'APPID_{ente}']
    client_secret =  os.environ[f'APPSECRET_{ente}']

    data = parse.urlencode({
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': 'https://graph.microsoft.com/.default',
        'grant_type': 'client_credentials'
    })

    uri = 'https://login.microsoftonline.com/{0}/oauth2/v2.0/token'.format(tenant_id)
    r = requests.post(uri, data=data).json()

    if not 'access_token' in r:
        print(f'{r}')
        exit(1)    
    return r['access_token']


def download_call_data(t, call_id):
    uri = f'https://graph.microsoft.com/beta/communications/callRecords/{call_id}?$expand=sessions($expand=segments)'
    head = { 'Authorization': f'Bearer {t}' }
    r = requests.get(uri, headers=head)
    response = r.json()

    if 'error' in response:
        print (f'{response}')
        return None
    else:
        return json.dumps(response)


def get_usernamefromid(t, userid, displayName=False):
    if userid == 'Sconosciuto': return 'Sconosciuto'

    keyname = userid
    username = f'Sconosciuto ({userid})'
    if displayName:
        keyname = f'{keyname}_d'

    uri = f'https://graph.microsoft.com/beta/users/{userid}'
    head = { 'Authorization': f'Bearer {t}' }
    r = requests.get(uri, headers=head)
    user = r.json()

    if displayName:
        if 'displayName' in user:
            username = user['displayName']
    else:   
        if 'surname' in user and 'givenName' in user:            
            username = f'{user["surname"]} {user["givenName"]}'

    return username


def save_virtual_workbook(workbook):
    with NamedTemporaryFile() as tf:
        workbook.save(tf.name)
        in_memory = BytesIO(tf.read())
        return in_memory.getvalue()
        

def sheet_registro(filename, participants, workbook):
    worksheet = workbook.active
    worksheet.title = 'Registro'
    worksheet.append(['Partecipante', 'Inizio presenza', 'Fine presenza', 'Tempo di partecipazione'])
    for cell in worksheet["1:1"]:
        cell.font = Font(bold=True)
    i = 1
    for pp in participants:
        i = i + 1
        start = datetime.strptime(pp['start'].split('.', 1)[0], '%Y-%m-%dT%H:%M:%S')
        end = datetime.strptime(pp['end'].split('.', 1)[0], '%Y-%m-%dT%H:%M:%S')
        delta = (end - start).seconds
        delta /= 60*60*24
        worksheet.append([
                pp['name'],
                start,
                end,
                delta
            ])
        cell = worksheet.cell(i, 2)
        cell.number_format = 'dd/mm/yyyy hh:mm'
        cell = worksheet.cell(i, 3)
        cell.number_format = 'dd/mm/yyyy hh:mm'
        cell = worksheet.cell(i, 4)
        cell.number_format = 'h "ore e" mm "minuti"'
    mediumStyle = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    worksheet.add_table(Table(ref=f'A1:D{i}', displayName='RegistroPresenze', tableStyleInfo=mediumStyle))
    worksheet.sheet_view.showGridLines = False
    report_id = os.path.basename(filename).replace('.json', '')
    worksheet.append([f''])
    worksheet.append([f'Report generato per il meeting con ID: {report_id}'])
    column_widths = [30, 20, 20, 20]
    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i+1)].width = column_width


def merge_intervals(temp_tuple):
    temp_tuple.sort(key=lambda interval: interval[0])
    merged = [temp_tuple[0]]
    for current in temp_tuple:
        previous = merged[-1]
        if current[0] <= previous[1]:
            previous[1] = max(previous[1], current[1])
        else:
            merged.append(current)
    return merged


def sheet_partecipazione(filename, start_time, participants, workbook):
    worksheet = workbook.create_sheet('Partecipazione')
    titles = ['Partecipante']
    max_interruptions = 0
    for pp in participants:
        if len(pp['periods']) > max_interruptions:
            max_interruptions = len(pp['periods'])

    for i in range(max_interruptions):
        titles.append(f'Inizio {i+1}')
        titles.append(f'Fine {i+1}') 

    worksheet.append(titles)
    for cell in worksheet["1:1"]:
        cell.font = Font(bold=True)
    
    i = 1
    for pp in participants:
        i = i + 1

        cur_start = start_time
        row = [pp['name']]
        for j in range(len(pp['periods'])):
            p = pp['periods'][j]
            row.append(p[0] - cur_start)
            row.append(p[1] - p[0])
            cur_start = p[1]

        worksheet.append(row)
        for ii in range(1, len(row)+1):
            cell = worksheet.cell(i, ii)
            cell.number_format = 'hh "ore e" mm "minuti"'

    column_widths = [30]
    for _i in range(max_interruptions*2):
        column_widths.append(20)
    for ii, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(ii+1)].width = column_width
        
    max_col = get_column_letter((max_interruptions*2)+1)
    mediumStyle = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    worksheet.add_table(Table(ref=f'A1:{max_col}{i}', displayName='RegistroPartecipazione', tableStyleInfo=mediumStyle))
    worksheet.sheet_view.showGridLines = False
    report_id = os.path.basename(filename).replace('.json', '')
    worksheet.append([f''])
    worksheet.append([f'Report generato per il meeting con ID: {report_id}'])

    data = Reference(worksheet, min_col=2, min_row=1, max_row=i, max_col=(max_interruptions*2)+1)
    cats = Reference(worksheet, min_col=1, min_row=2, max_row=i)
    chart = BarChart()
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.legend = None 
    chart.height = 20
    chart.width = 40
    chart.x_axis.scaling.orientation = "maxMin"
    for ii, s in enumerate(chart.series):
        s.graphicalProperties.line.nofill = True
        s.graphicalProperties.line.shadow = None

        if ii % 2 == 0:
            s.graphicalProperties.noFill = True
        else:
            s.graphicalProperties.solidFill = "008F00"

    worksheet.add_chart(chart, "A43")


def download_generatedexcel(t, jsonFileData, report_id=None):
    name = 'Sconosciuto'
    if 'organizer' in jsonFileData and jsonFileData['organizer'] is not None:
        if 'user' in jsonFileData['organizer'] and jsonFileData['organizer']['user'] is not None:
            if 'id' in jsonFileData['organizer']['user']:
                name = get_usernamefromid(t, jsonFileData['organizer']['user']['id'], True)

    start_time = datetime.strptime(jsonFileData['startDateTime'].split('.', 1)[0].split('Z', 1)[0], '%Y-%m-%dT%H:%M:%S')
    start_time = start_time.replace(tzinfo=from_zone).astimezone(to_zone)
    start_time_name = start_time.strftime('%Y-%m-%d %H-%M')
    dest_filename = f'{start_time_name} - {name}.xlsx'

    users = {}
    for c in jsonFileData['sessions']:
        if 'caller' not in c or c['caller'] is None:
            continue
        if 'identity' not in c['caller'] or c['caller']['identity'] is None:
            continue
        if 'user' not in c['caller']['identity'] or c['caller']['identity']['user'] is None:
            continue
        if 'id' not in c['caller']['identity']['user'] or c['caller']['identity']['user']['id'] is None:
            continue

        if 'userAgent' in c['caller'] and 'headerValue' in c['caller']['userAgent'] and c['caller']['userAgent']['headerValue'] == 'SkypeBot Call Recorder Teams':
            continue

        curuid = c['caller']['identity']['user']['id']
        if curuid not in users:
            displayname = get_usernamefromid(t, curuid)
            users[curuid] = { 'name': displayname, 'min_start': None, 'max_end': None, 'duration': 0 }

        if 'periods' not in users[curuid]: users[curuid]['periods'] = []

        start = datetime.strptime(c['startDateTime'].split('.', 1)[0].split('Z', 1)[0], '%Y-%m-%dT%H:%M:%S')
        start = start.replace(tzinfo=from_zone).astimezone(to_zone)
        if users[curuid]['min_start'] is None or start < users[curuid]['min_start']:
            users[curuid]['min_start'] = start

        end = datetime.strptime(c['endDateTime'].split('.', 1)[0].split('Z', 1)[0], '%Y-%m-%dT%H:%M:%S')
        end = end.replace(tzinfo=from_zone).astimezone(to_zone)
        if users[curuid]['max_end'] is None or end > users[curuid]['max_end']:
            users[curuid]['max_end'] = end
        
        if end is not None and start is not None:
            delta = end - start
            users[curuid]['duration'] += delta.seconds

        users[curuid]['periods'].append([start, end])

    participants = []
    for _uid, data in users.items():
        data['duration'] /= 60
        
        hours = math.floor(data['duration'] / 60)
        minutes = math.floor(data['duration'] % 60)
        duration = "{0} ore e {1} minuti".format(hours, minutes)

        participants.append({
            'uid': _uid,
            'name': data['name'],
            'start': data['min_start'].strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
            'end': data['max_end'].strftime('%Y-%m-%dT%H:%M:%S.%fZ'),
            'duration': duration,
            'periods': merge_intervals(data['periods'])
        })

    if len(participants) <= 1:
        return dest_filename, None

    participants.sort(key=lambda stud: 'zzz_{}'.format(stud['name']) if 'Sconosciuto' in stud['name'] else stud['name'] )

    ###########

    workbook = Workbook()
    sheet_registro(report_id, participants, workbook)
    sheet_partecipazione(report_id, start_time, participants, workbook)

    return dest_filename, save_virtual_workbook(workbook)